"""Excel export with openpyxl formatting, headers, and merged cells."""
from __future__ import annotations

from pathlib import Path

import pandas as pd

from .lesson_plan_format import (
    LAB_MAIN_COLUMNS,
    THEORY_MAIN_COLUMNS,
    formatted_lesson_plan_frame,
    lab_date_row_spans,
    summary_values,
    week_row_spans,
)
from .utils import LessonPlanBundle, escape_cell_value, header_image_path, report_output_paths

try:
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.styles import Alignment, Border, Font, Side
except ImportError:  # pragma: no cover - depends on local environment
    OpenpyxlImage = None
    Alignment = None
    Border = None
    Font = None
    Side = None


HEADER_START_ROW = 7
FORMATTED_HEADER_ROW = 18
FORMATTED_DATA_ROW = 19
FREEZE_PANE = 'A19'
SIGNATURE_ROW_HEIGHT = 30
SIGNATURE_TOP_SPACER_ROW_HEIGHT = 28
SIGNATURE_BOTTOM_SPACER_ROW_HEIGHT = 18
SIGNATURE_TOP_SPACER_ROWS = 2


def export_bundle_to_excel(bundle: LessonPlanBundle, output_dir: str | Path) -> dict[str, Path]:
    output_paths = report_output_paths(bundle, output_dir)
    lesson_plan_path = output_paths['lesson_plan_excel']
    monthly_plan_path = output_paths['monthly_plan_excel']
    coverage_path = output_paths['coverage_report_excel']

    with pd.ExcelWriter(lesson_plan_path, engine='openpyxl') as writer:
        workbook = writer.book
        lesson_sheet = workbook.create_sheet('Lesson Plan', 0)
        if summary_values(bundle)['is_lab']:
            _build_formatted_lab_sheet(lesson_sheet, bundle)
        else:
            _build_formatted_theory_sheet(lesson_sheet, bundle)
        _write_frame_sheet(writer, 'Monthly Plan', bundle.monthly_plan)
        _write_frame_sheet(writer, 'Coverage Report', bundle.coverage_report)

        for worksheet in workbook.worksheets:
            _insert_header_image(worksheet)
            if worksheet.title == 'Lesson Plan':
                worksheet.freeze_panes = FREEZE_PANE
            else:
                worksheet.freeze_panes = 'A9'
                _autosize_worksheet(worksheet)

    _write_workbook_with_header(monthly_plan_path, {'Monthly Plan': bundle.monthly_plan})
    _write_workbook_with_header(coverage_path, {'Coverage Report': bundle.coverage_report})

    return {
        'lesson_plan': lesson_plan_path,
        'monthly_plan': monthly_plan_path,
        'coverage_report': coverage_path,
    }


def _write_workbook_with_header(workbook_path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(workbook_path, engine='openpyxl') as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=HEADER_START_ROW)

        workbook = writer.book
        for worksheet in workbook.worksheets:
            _insert_header_image(worksheet)
            worksheet.freeze_panes = 'A9'
            _autosize_worksheet(worksheet)


def _write_frame_sheet(writer, sheet_name: str, frame: pd.DataFrame) -> None:
    frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=HEADER_START_ROW)


def _build_formatted_theory_sheet(worksheet, bundle: LessonPlanBundle) -> None:
    summary = summary_values(bundle)
    frame = formatted_lesson_plan_frame(bundle.lesson_plan, lab_mode=False)
    sign_row = 16
    regulation_row = sign_row + 1

    _write_common_heading(worksheet, summary)
    _write_summary_block(worksheet, summary)
    _write_signature_row(worksheet, sign_row, summary['sign_labels'], total_columns=9)
    worksheet.merge_cells(start_row=regulation_row, start_column=1, end_row=regulation_row, end_column=9)
    worksheet.cell(row=regulation_row, column=1, value=f"Regulation: {summary['regulation']}     Academic Year: {summary['academic_year']}")

    for column_index, column_name in enumerate(THEORY_MAIN_COLUMNS, start=1):
        worksheet.cell(row=FORMATTED_HEADER_ROW, column=column_index, value=column_name)

    for row_index, row_values in enumerate(frame.values.tolist(), start=FORMATTED_DATA_ROW):
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=escape_cell_value(value))

    for start_index, end_index, label in week_row_spans(frame):
        start_row = FORMATTED_DATA_ROW + start_index
        end_row = FORMATTED_DATA_ROW + end_index
        worksheet.cell(row=start_row, column=2, value=label)
        if end_row > start_row:
            worksheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)

    _style_theory_sheet(worksheet, len(frame), sign_row, regulation_row)


def _build_formatted_lab_sheet(worksheet, bundle: LessonPlanBundle) -> None:
    summary = summary_values(bundle)
    frame = formatted_lesson_plan_frame(bundle.lesson_plan, lab_mode=True)

    _write_common_heading(worksheet, summary)
    _write_summary_block(worksheet, summary)
    sign_row = FORMATTED_DATA_ROW + len(frame) + SIGNATURE_TOP_SPACER_ROWS + 2
    _write_signature_row(worksheet, sign_row, summary['sign_labels'], total_columns=7)

    for column_index, column_name in enumerate(LAB_MAIN_COLUMNS, start=1):
        worksheet.cell(row=FORMATTED_HEADER_ROW, column=column_index, value=column_name)

    for row_index, row_values in enumerate(frame.values.tolist(), start=FORMATTED_DATA_ROW):
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=escape_cell_value(value))

    for start_index, end_index, _label in lab_date_row_spans(frame):
        start_row = FORMATTED_DATA_ROW + start_index
        end_row = FORMATTED_DATA_ROW + end_index
        if end_row > start_row:
            worksheet.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5)

    _style_lab_sheet(worksheet, len(frame), sign_row)


def _write_common_heading(worksheet, summary: dict[str, str]) -> None:
    worksheet.merge_cells('A7:I7')
    worksheet['A7'] = summary['department_name']
    worksheet.merge_cells('A9:I9')
    worksheet['A9'] = summary['plan_title']


def _write_summary_block(worksheet, summary: dict[str, str]) -> None:
    worksheet['A11'] = 'Subject\nCode'
    worksheet['B11'] = 'Subject Name'
    worksheet['C11'] = 'Class/Sem'
    worksheet['D11'] = 'Faculty Name /\nDesignation'
    worksheet['E11'] = 'Number of\nStudents'
    worksheet.merge_cells('F11:G11')
    worksheet['F11'] = summary['total_label']

    worksheet['A12'] = summary['course_code']
    worksheet['B12'] = summary['course_title_display']
    worksheet['C12'] = summary['class_sem_display']
    worksheet['D12'] = summary['faculty_display_summary']
    worksheet['E12'] = summary['student_count']
    worksheet['F12'] = summary['lecture_label']
    worksheet['G12'] = summary['tutorial_label']

    worksheet['A13'] = summary['course_code']
    worksheet['B13'] = summary['course_title_display']
    worksheet['C13'] = summary['class_sem_display']
    worksheet['D13'] = summary['faculty_display_summary']
    worksheet['E13'] = summary['student_count']
    worksheet['F13'] = summary['lecture_count']
    worksheet['G13'] = summary['tutorial_count']

    worksheet.merge_cells('A12:A13')
    worksheet.merge_cells('B12:B13')
    worksheet.merge_cells('C12:C13')
    worksheet.merge_cells('D12:D13')
    worksheet.merge_cells('E12:E13')


def _write_signature_row(worksheet, row: int, labels: tuple[str, str, str], total_columns: int) -> None:
    for start_column, end_column, label in _signature_ranges(total_columns, labels):
        if end_column > start_column:
            worksheet.merge_cells(start_row=row, start_column=start_column, end_row=row, end_column=end_column)
        worksheet.cell(row=row, column=start_column, value=label)


def _signature_ranges(total_columns: int, labels: tuple[str, str, str]) -> list[tuple[int, int, str]]:
    if total_columns >= 9:
        return [(1, 2, labels[0]), (5, 5, labels[1]), (8, 9, labels[2])]
    if total_columns >= 7:
        return [(1, 2, labels[0]), (4, 4, labels[1]), (6, 7, labels[2])]
    midpoint = max(2, total_columns // 2)
    return [(1, 1, labels[0]), (midpoint, midpoint, labels[1]), (total_columns, total_columns, labels[2])]


def _style_theory_sheet(worksheet, data_length: int, sign_row: int, regulation_row: int) -> None:
    if Font is None or Alignment is None or Border is None or Side is None:
        return

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name='Times New Roman', size=12, bold=True)
    subtitle_font = Font(name='Times New Roman', size=14, bold=True)
    header_font = Font(name='Times New Roman', size=9, bold=True)
    cell_font = Font(name='Times New Roman', size=9)

    _style_common_headings(worksheet, title_font, subtitle_font)
    _style_signature_row(worksheet, sign_row, border, total_columns=9)
    worksheet.cell(row=regulation_row, column=1).font = Font(name='Times New Roman', size=9, bold=True)
    worksheet.cell(row=regulation_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    _style_summary_block(worksheet, border, header_font, cell_font)

    for column in range(1, 10):
        cell = worksheet.cell(row=FORMATTED_HEADER_ROW, column=column)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = header_font

    for row in range(FORMATTED_DATA_ROW, FORMATTED_DATA_ROW + data_length):
        for column in range(1, 10):
            cell = worksheet.cell(row=row, column=column)
            cell.border = border
            cell.font = cell_font
            cell.alignment = Alignment(horizontal='left' if column == 4 else 'center', vertical='center', wrap_text=True)

    for letter, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'], [8, 10, 11, 50, 14, 14, 11, 22, 16]):
        worksheet.column_dimensions[letter].width = width


def _style_lab_sheet(worksheet, data_length: int, sign_row: int) -> None:
    if Font is None or Alignment is None or Border is None or Side is None:
        return

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name='Times New Roman', size=12, bold=True)
    subtitle_font = Font(name='Times New Roman', size=14, bold=True)
    header_font = Font(name='Times New Roman', size=9, bold=True)
    cell_font = Font(name='Times New Roman', size=9)

    _style_common_headings(worksheet, title_font, subtitle_font)
    _style_summary_block(worksheet, border, header_font, cell_font)
    _style_signature_row(worksheet, sign_row, border, total_columns=7)

    for column in range(1, 8):
        cell = worksheet.cell(row=FORMATTED_HEADER_ROW, column=column)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = header_font

    for row in range(FORMATTED_DATA_ROW, FORMATTED_DATA_ROW + data_length):
        for column in range(1, 8):
            cell = worksheet.cell(row=row, column=column)
            cell.border = border
            cell.font = cell_font
            cell.alignment = Alignment(horizontal='left' if column == 3 else 'center', vertical='center', wrap_text=True)

    for letter, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G'], [6, 16, 42, 8, 12, 12, 12]):
        worksheet.column_dimensions[letter].width = width


def _style_common_headings(worksheet, title_font, subtitle_font) -> None:
    worksheet['A7'].font = title_font
    worksheet['A7'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['A9'].font = subtitle_font
    worksheet['A9'].alignment = Alignment(horizontal='center', vertical='center')


def _style_summary_block(worksheet, border, header_font, cell_font) -> None:
    for row in range(11, 14):
        for column in range(1, 8):
            cell = worksheet.cell(row=row, column=column)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = header_font if row == 11 else cell_font
    worksheet.row_dimensions[11].height = 30
    worksheet.row_dimensions[12].height = 38
    worksheet.row_dimensions[13].height = 28


def _style_signature_row(worksheet, row: int, border, total_columns: int) -> None:
    signature_font = Font(name='Times New Roman', size=10, bold=True)
    for column in range(1, total_columns + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.border = Border()
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if cell.value:
            cell.font = signature_font
    worksheet.row_dimensions[row].height = SIGNATURE_ROW_HEIGHT
    for offset in range(1, SIGNATURE_TOP_SPACER_ROWS + 1):
        spacer_row = row - offset
        if spacer_row > 1:
            worksheet.row_dimensions[spacer_row].height = SIGNATURE_TOP_SPACER_ROW_HEIGHT
    worksheet.row_dimensions[row + 1].height = SIGNATURE_BOTTOM_SPACER_ROW_HEIGHT


def _insert_header_image(worksheet) -> None:
    image_path = header_image_path()
    if image_path is None or OpenpyxlImage is None:
        return

    image = OpenpyxlImage(str(image_path))
    image.width = 620
    image.height = 98
    worksheet.add_image(image, 'A1')
    worksheet.row_dimensions[1].height = 76


def _autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
